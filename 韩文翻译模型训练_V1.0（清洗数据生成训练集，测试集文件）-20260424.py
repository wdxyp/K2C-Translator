import re
import numpy as np
from sklearn.model_selection import train_test_split
import torch
import torch.nn as nn
import torch.optim as optim
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import os
import openpyxl

# 选择语料库文件（XLSX格式，B列为韩文，D列为中文）
def select_corpus_files():
    root = tk.Tk()
    root.withdraw()

    print("请选择语料库XLSX文件（可多选）")
    corpus_file_paths = filedialog.askopenfilenames(title="选择语料库文件", filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")])

    return corpus_file_paths

# 从XLSX文件读取语料库，B列为韩文，D列为中文
def read_corpus(corpus_file_paths):
    all_korean_sentences = []
    all_chinese_sentences = []

    for corpus_file_path in corpus_file_paths:
        try:
            wb = openpyxl.load_workbook(corpus_file_path, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 4:
                    korean_sentence = str(row[1]).strip() if row[1] else ""
                    chinese_sentence = str(row[3]).strip() if row[3] else ""
                    if korean_sentence and chinese_sentence:
                        all_korean_sentences.append(korean_sentence)
                        all_chinese_sentences.append(chinese_sentence)

            wb.close()
        except Exception as e:
            print(f"读取文件 {corpus_file_path} 时出错: {e}")

    if len(all_korean_sentences) == 0 or len(all_chinese_sentences) == 0:
        print("未读取到任何句子，请检查文件内容。")

    return all_korean_sentences, all_chinese_sentences

# 清洗文本，去除特殊字符
def clean_text(sentence):
    sentence = re.sub(r'[^\w\s]', '', sentence)
    return sentence.strip()

# 分词（对于韩语和中文，可使用不同的分词工具）
def tokenize(sentences):
    tokenized = []
    for sentence in sentences:
        tokens = sentence.split()  # 简单示例，实际可能需要更复杂的分词方法
        tokenized.append(tokens)
    return tokenized

# 创建交互界面
def create_gui():
    root = tk.Tk()
    root.title("语料库文件选择和保存")

    global corpus_file_paths
    corpus_file_paths = []

    file_text = tk.Text(root, height=5, width=50)
    file_text.pack(pady=5)

    def select_files():
        global corpus_file_paths
        corpus_file_paths = filedialog.askopenfilenames(title="选择语料库XLSX文件", filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")])

        file_text.delete(1.0, tk.END)
        for path in corpus_file_paths:
            file_text.insert(tk.END, path + "\n")

        status_label.config(text=f"已选择 {len(corpus_file_paths)} 个文件")

    def save_files():
        global corpus_file_paths
        if not corpus_file_paths:
            status_label.config(text="请先选择语料库文件")
            return

        korean_sentences, chinese_sentences = read_corpus(corpus_file_paths)

        if not korean_sentences or not chinese_sentences:
            status_label.config(text="未读取到有效句子，请检查XLSX文件格式")
            return

        non_empty_indices = [i for i, (ko, zh) in enumerate(zip(korean_sentences, chinese_sentences)) if ko and zh]
        korean_sentences = [korean_sentences[i] for i in non_empty_indices]
        chinese_sentences = [chinese_sentences[i] for i in non_empty_indices]

        korean_sentences = [clean_text(sent) for sent in korean_sentences]
        chinese_sentences = [clean_text(sent) for sent in chinese_sentences]

        korean_tokens = tokenize(korean_sentences)
        chinese_tokens = tokenize(chinese_sentences)

        ko_train, ko_test, zh_train, zh_test = train_test_split(korean_tokens, chinese_tokens, test_size=0.2, random_state=42)

        def tokens_to_string(tokens_list):
            return [' '.join(tokens) for tokens in tokens_list]

        ko_train_str = tokens_to_string(ko_train)
        ko_test_str = tokens_to_string(ko_test)
        zh_train_str = tokens_to_string(zh_train)
        zh_test_str = tokens_to_string(zh_test)

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        save_folder = "Training Data"

        if not os.path.exists(save_folder):
            os.makedirs(save_folder)

        train_filename = os.path.join(save_folder, f"train_{timestamp}.xlsx")
        test_filename = os.path.join(save_folder, f"test_{timestamp}.xlsx")

        train_wb = openpyxl.Workbook()
        train_ws = train_wb.active
        train_ws.append(['韩语', '中文'])
        for ko, zh in zip(ko_train_str, zh_train_str):
            if ko and zh:
                train_ws.append([ko, zh])
        train_wb.save(train_filename)
        train_wb.close()

        test_wb = openpyxl.Workbook()
        test_ws = test_wb.active
        test_ws.append(['韩语', '中文'])
        for ko, zh in zip(ko_test_str, zh_test_str):
            if ko and zh:
                test_ws.append([ko, zh])
        test_wb.save(test_filename)
        test_wb.close()

        status_label.config(text=f"训练集: {train_filename}\n测试集: {test_filename}")

    tk.Button(root, text="选择语料库文件", command=select_files).pack(pady=5)
    tk.Button(root, text="开始处理", command=save_files).pack(pady=5)

    status_label = tk.Label(root, text="")
    status_label.pack(pady=10)

    root.mainloop()

# 运行交互界面
create_gui()
