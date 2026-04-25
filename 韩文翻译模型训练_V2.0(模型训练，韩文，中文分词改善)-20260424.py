import re
import numpy as np
from sklearn.model_selection import train_test_split
import torch
import torch.nn as nn
import torch.optim as optim
from collections import Counter
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import os
import pickle
import jieba
import openpyxl
from konlpy.tag import Okt

# 选择语料库文件（XLSX格式，B列为韩文，D列为中文）
def select_corpus_files():
    root = tk.Tk()
    root.withdraw()

    print("请选择语料库XLSX文件（可多选）")
    corpus_file_paths = filedialog.askopenfilenames(title="选择语料库文件", filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")])

    return corpus_file_paths

# 从XLSX文件读取语料库，B列为韩文，D列为中文
def read_corpus(corpus_file_paths):
    all_korean_sentences = []
    all_chinese_sentences = []

    for corpus_file_path in corpus_file_paths:
        try:
            wb = openpyxl.load_workbook(corpus_file_path, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 4:
                    korean_sentence = str(row[1]).strip() if row[1] else ""
                    chinese_sentence = str(row[3]).strip() if row[3] else ""
                    if korean_sentence and chinese_sentence:
                        all_korean_sentences.append(korean_sentence)
                        all_chinese_sentences.append(chinese_sentence)

            wb.close()
        except Exception as e:
            print(f"读取文件 {corpus_file_path} 时出错: {e}")

    if len(all_korean_sentences) == 0 or len(all_chinese_sentences) == 0:
        print("未读取到任何句子，请检查文件内容。")

    return all_korean_sentences, all_chinese_sentences

# 清洗文本，去除特殊字符
def clean_text(sentence):
    sentence = re.sub(r'[^\w\s]', '', sentence)
    return sentence.strip()

# 分词（对于韩语和中文，使用不同的分词工具）
def tokenize(sentences, lang):
    tokenized = []
    total = len(sentences)
    print(f"正在对 {lang} 语料进行分词，总计 {total} 条句子...")
    if lang == 'ko':
        try:
            okt = Okt()
            for i, sentence in enumerate(sentences):
                tokens = okt.morphs(sentence)
                tokenized.append(tokens)
                if (i + 1) % 1000 == 0:
                    print(f"  已完成 {i+1}/{total} ({(i+1)/total*100:.1f}%)", end='\r')
        except Exception as e:
            print(f"韩文分词器启动失败（可能是未安装 Java 或环境变量未配置）: {e}")
            print("尝试使用简单空格分词作为备选方案...")
            for i, sentence in enumerate(sentences):
                tokens = sentence.split()
                tokenized.append(tokens)
                if (i + 1) % 1000 == 0:
                    print(f"  已完成 {i+1}/{total} ({(i+1)/total*100:.1f}%)", end='\r')
    elif lang == 'zh':
        for i, sentence in enumerate(sentences):
            tokens = jieba.lcut(sentence) # jieba 不需要 Java，分词很准
            tokenized.append(tokens)
            if (i + 1) % 1000 == 0:
                print(f"  已完成 {i+1}/{total} ({(i+1)/total*100:.1f}%)", end='\r')
    print(f"\n{lang} 分词完成！")
    return tokenized

# 创建交互界面
def create_gui():
    root = tk.Tk()
    root.title("语料库文件选择和保存")

    # 全局变量用于存储文件路径
    global corpus_file_paths
    corpus_file_paths = []

    # 显示文件路径的文本框
    file_text = tk.Text(root, height=5, width=50)
    file_text.grid(row=0, column=0, columnspan=2, pady=5)

    # 选择语料库文件
    def select_files():
        global corpus_file_paths
        corpus_file_paths = filedialog.askopenfilenames(title="选择语料库XLSX文件", filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")])

        # 清空文本框
        file_text.delete(1.0, tk.END)

        # 显示文件路径
        for path in corpus_file_paths:
            file_text.insert(tk.END, path + "\n")

        status_label.config(text=f"已选择 {len(corpus_file_paths)} 个文件")

    # 保存文件
    def save_files():
        global corpus_file_paths
        if not corpus_file_paths:
            status_label.config(text="请先选择语料库文件")
            return

        korean_sentences, chinese_sentences = read_corpus(corpus_file_paths)

        if not korean_sentences or not chinese_sentences:
            status_label.config(text="未读取到有效句子，请检查XLSX文件格式")
            return

        # 过滤掉空句子
        non_empty_indices = [i for i, (ko, zh) in enumerate(zip(korean_sentences, chinese_sentences)) if ko and zh]
        korean_sentences = [korean_sentences[i] for i in non_empty_indices]
        chinese_sentences = [chinese_sentences[i] for i in non_empty_indices]

        # 清洗文本
        korean_sentences = [clean_text(sent) for sent in korean_sentences]
        chinese_sentences = [clean_text(sent) for sent in chinese_sentences]

        # 分词
        korean_tokens = tokenize(korean_sentences, 'ko')
        chinese_tokens = tokenize(chinese_sentences, 'zh')

        # 划分训练集和测试集
        ko_train, ko_test, zh_train, zh_test = train_test_split(korean_tokens, chinese_tokens, test_size=0.2, random_state=42)

        # 将分词后的列表转换为字符串
        def tokens_to_string(tokens_list):
            return [' '.join(tokens) for tokens in tokens_list]

        ko_train_str = tokens_to_string(ko_train)
        ko_test_str = tokens_to_string(ko_test)
        zh_train_str = tokens_to_string(zh_train)
        zh_test_str = tokens_to_string(zh_test)

        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        # 定义保存文件夹名
        save_folder = f"Training Data"

        # 检查文件夹是否存在，如果不存在则创建
        if not os.path.exists(save_folder):
            os.makedirs(save_folder)

        train_filename = os.path.join(save_folder, f"train_{timestamp}.xlsx")
        test_filename = os.path.join(save_folder, f"test_{timestamp}.xlsx")

        # 保存训练集到 XLSX 文件
        train_wb = openpyxl.Workbook()
        train_ws = train_wb.active
        train_ws.append(['韩语', '中文'])
        for ko, zh in zip(ko_train_str, zh_train_str):
            if ko and zh:
                train_ws.append([ko, zh])
        train_wb.save(train_filename)
        train_wb.close()

        # 保存测试集到 XLSX 文件
        test_wb = openpyxl.Workbook()
        test_ws = test_wb.active
        test_ws.append(['韩语', '中文'])
        for ko, zh in zip(ko_test_str, zh_test_str):
            if ko and zh:
                test_ws.append([ko, zh])
        test_wb.save(test_filename)
        test_wb.close()

        status_label.config(text=f"训练集已保存到 {train_filename}\n测试集已保存到 {test_filename}")
        # 训练模型
        train_model(train_filename, test_filename)

    # 预览词汇表
    def preview_vocab():
        global corpus_file_paths
        if not corpus_file_paths:
            status_label.config(text="请先选择语料库文件")
            return

        korean_sentences, chinese_sentences = read_corpus(corpus_file_paths)

        # 过滤掉空句子
        non_empty_indices = [i for i, (ko, zh) in enumerate(zip(korean_sentences, chinese_sentences)) if ko and zh]
        korean_sentences = [korean_sentences[i] for i in non_empty_indices]
        chinese_sentences = [chinese_sentences[i] for i in non_empty_indices]

        # 清洗文本
        korean_sentences = [clean_text(sent) for sent in korean_sentences]
        chinese_sentences = [clean_text(sent) for sent in chinese_sentences]

        # 分词
        korean_tokens = tokenize(korean_sentences, 'ko')
        chinese_tokens = tokenize(chinese_sentences, 'zh')

        # 构建词汇表
        korean_vocab = build_vocab(korean_tokens)
        chinese_vocab = build_vocab(chinese_tokens)

        # 清空文本框
        korean_vocab_text.delete(1.0, tk.END)
        chinese_vocab_text.delete(1.0, tk.END)

        # 显示韩语词汇表
        korean_vocab_text.insert(tk.END, "韩语词汇表:\n")
        for word, index in korean_vocab.items():
            korean_vocab_text.insert(tk.END, f"{word}: {index}\n")

        # 显示中文词汇表
        chinese_vocab_text.insert(tk.END, "中文词汇表:\n")
        for word, index in chinese_vocab.items():
            chinese_vocab_text.insert(tk.END, f"{word}: {index}\n")

    # 开始按钮
    start_button = tk.Button(root, text="开始", command=save_files)
    start_button.grid(row=2, column=0, columnspan=2, pady=10)

    # 停止按钮（目前只是占位，可根据需求添加停止逻辑）
    stop_button = tk.Button(root, text="停止", command=lambda: status_label.config(text="停止操作未实现"))
    stop_button.grid(row=3, column=0, columnspan=2, pady=10)

    # 选择文件按钮
    select_button = tk.Button(root, text="选择语料库文件", command=select_files)
    select_button.grid(row=4, column=0, columnspan=2, pady=10)

    # 预览词汇表按钮
    preview_button = tk.Button(root, text="预览词汇表", command=preview_vocab)
    preview_button.grid(row=5, column=0, columnspan=2, pady=10)

    # 显示韩语词汇表的文本框
    korean_vocab_text = tk.Text(root, height=10, width=50)
    korean_vocab_text.grid(row=6, column=0, pady=5, padx=5)

    # 显示中文词汇表的文本框
    chinese_vocab_text = tk.Text(root, height=10, width=50)
    chinese_vocab_text.grid(row=6, column=1, pady=5, padx=5)

    status_label = tk.Label(root, text="")
    status_label.grid(row=7, column=0, columnspan=2, pady=10)

    root.mainloop()

# 加载数据集
def load_dataset(file_path):
    korean_sentences = []
    chinese_sentences = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2:
                ko_sent = str(row[0]).split() if row[0] else []
                zh_sent = str(row[1]).split() if row[1] else []
                if ko_sent and zh_sent:
                    korean_sentences.append(ko_sent)
                    chinese_sentences.append(zh_sent)
        wb.close()
    except Exception as e:
        print(f"加载数据集 {file_path} 时出错: {e}")
    return korean_sentences, chinese_sentences

# 构建词汇表，支持最小频率过滤和最大容量限制
def build_vocab(sentences, min_freq=2, max_size=30000):
    counter = Counter()
    for sentence in sentences:
        for word in sentence:
            if isinstance(word, str) and word.strip():
                counter[word] += 1
    
    # 按照频率排序，只取前 max_size 个
    most_common = counter.most_common(max_size)
    
    vocab = {'<pad>': 0, '<sos>': 1, '<eos>': 2, '<unk>': 3}
    index = 4
    for word, freq in most_common:
        if freq >= min_freq:
            if word not in vocab:
                vocab[word] = index
                index += 1
    return vocab

# 文本转张量，支持 <unk> 机制
def text_to_tensor(sentences, vocab):
    tensors = []
    unk_idx = vocab.get('<unk>', 0) # 如果没有 <unk> 则退回到 <pad>
    for sentence in sentences:
        tensor = [vocab['<sos>']] + [vocab.get(word, unk_idx) for word in sentence] + [vocab['<eos>']]
        tensors.append(torch.LongTensor(tensor))
    return tensors

# 定义编码器
class Encoder(nn.Module):
    def __init__(self, input_dim, emb_dim, hid_dim, n_layers, dropout):
        super().__init__()
        self.hid_dim = hid_dim
        self.n_layers = n_layers
        self.embedding = nn.Embedding(input_dim, emb_dim)
        self.rnn = nn.LSTM(emb_dim, hid_dim, n_layers, dropout=dropout)
        self.dropout = nn.Dropout(dropout)

    def forward(self, src, src_lens):
        embedded = self.dropout(self.embedding(src))
        packed = torch.nn.utils.rnn.pack_padded_sequence(embedded, src_lens, batch_first=False, enforce_sorted=False)
        outputs, (hidden, cell) = self.rnn(packed)
        return hidden, cell

# 定义解码器
class Decoder(nn.Module):
    def __init__(self, output_dim, emb_dim, hid_dim, n_layers, dropout):
        super().__init__()
        self.output_dim = output_dim
        self.hid_dim = hid_dim
        self.n_layers = n_layers
        self.embedding = nn.Embedding(output_dim, emb_dim)
        self.rnn = nn.LSTM(emb_dim, hid_dim, n_layers, dropout=dropout)
        self.fc_out = nn.Linear(hid_dim, output_dim)
        self.dropout = nn.Dropout(dropout)

    def forward(self, input, hidden, cell):
        input = input.unsqueeze(0)
        embedded = self.dropout(self.embedding(input))
        output, (hidden, cell) = self.rnn(embedded, (hidden, cell))
        prediction = self.fc_out(output.squeeze(0))
        return prediction, hidden, cell

# 定义Seq2Seq模型
class Seq2Seq(nn.Module):
    def __init__(self, encoder, decoder, device):
        super().__init__()
        self.encoder = encoder
        self.decoder = decoder
        self.device = device

    def forward(self, src, src_lens, trg, teacher_forcing_ratio=0.5):
        batch_size = trg.shape[1]
        trg_len = trg.shape[0]
        trg_vocab_size = self.decoder.output_dim
        outputs = torch.zeros(trg_len, batch_size, trg_vocab_size).to(self.device)
        hidden, cell = self.encoder(src, src_lens)
        input = trg[0, :]
        for t in range(1, trg_len):
            output, hidden, cell = self.decoder(input, hidden, cell)
            outputs[t] = output
            teacher_force = np.random.random() < teacher_forcing_ratio
            top1 = output.argmax(1)
            input = trg[t] if teacher_force else top1
        return outputs

def collate_fn(batch, pad_idx):
    ko_batch, zh_batch = zip(*batch)
    ko_lens = [len(seq) for seq in ko_batch]
    zh_lens = [len(seq) for seq in zh_batch]
    ko_padded = torch.nn.utils.rnn.pad_sequence(ko_batch, batch_first=True, padding_value=pad_idx)
    zh_padded = torch.nn.utils.rnn.pad_sequence(zh_batch, batch_first=True, padding_value=pad_idx)
    return ko_padded, ko_lens, zh_padded, zh_lens

# 训练模型
def train_model(train_file, test_file):
    ko_train, zh_train = load_dataset(train_file)
    ko_test, zh_test = load_dataset(test_file)

    # 设置词汇表参数
    MIN_FREQ = 2
    MAX_VOCAB_SIZE = 30000

    korean_vocab = build_vocab(ko_train, min_freq=MIN_FREQ, max_size=MAX_VOCAB_SIZE)
    chinese_vocab = build_vocab(zh_train, min_freq=MIN_FREQ, max_size=MAX_VOCAB_SIZE)
    vocab_reverse_chinese = {v: k for k, v in chinese_vocab.items()}

    ko_train_tensor = text_to_tensor(ko_train, korean_vocab)
    zh_train_tensor = text_to_tensor(zh_train, chinese_vocab)
    ko_test_tensor = text_to_tensor(ko_test, korean_vocab)
    zh_test_tensor = text_to_tensor(zh_test, chinese_vocab)

    print(f"\n数据集加载完成（已开启频率过滤：min_freq={MIN_FREQ}, max_size={MAX_VOCAB_SIZE}）：")
    print(f"  训练集句子数: {len(ko_train_tensor)}")
    print(f"  测试集句子数: {len(ko_test_tensor)}")
    print(f"  韩语词汇表大小: {len(korean_vocab)}")
    print(f"  中文词汇表大小: {len(chinese_vocab)}")

    INPUT_DIM = len(korean_vocab)
    OUTPUT_DIM = len(chinese_vocab)
    ENC_EMB_DIM = 256
    DEC_EMB_DIM = 256
    HID_DIM = 256
    N_LAYERS = 2
    ENC_DROPOUT = 0.4
    DEC_DROPOUT = 0.4
    BATCH_SIZE = 64
    N_EPOCHS = 50
    CLIP = 1
    PAD_IDX = korean_vocab['<pad>']
    EARLY_STOPPING_PATIENCE = 10
    WEIGHT_DECAY = 1e-5

    # 考虑到 128M 显存极低，强制使用 CPU 训练以利用系统内存
    device = torch.device('cpu')
    print(f"检测到显存较低 (128M)，已强制切换至 CPU 训练以确保程序稳定。")
    encoder = Encoder(INPUT_DIM, ENC_EMB_DIM, HID_DIM, N_LAYERS, ENC_DROPOUT)
    decoder = Decoder(OUTPUT_DIM, DEC_EMB_DIM, HID_DIM, N_LAYERS, DEC_DROPOUT)
    model = Seq2Seq(encoder, decoder, device).to(device)

    optimizer = optim.SGD(model.parameters(), lr=0.1, momentum=0.9, weight_decay=WEIGHT_DECAY)
    scheduler = optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='min', factor=0.5, patience=2, verbose=True)
    criterion = nn.CrossEntropyLoss(ignore_index=PAD_IDX)

    best_test_loss = float('inf')
    patience_counter = 0
    best_model_state = None
    train_losses = []
    test_losses = []

    train_dataset = list(zip(ko_train_tensor, zh_train_tensor))
    test_dataset = list(zip(ko_test_tensor, zh_test_tensor))
    train_loader = torch.utils.data.DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True, collate_fn=lambda b: collate_fn(b, PAD_IDX))
    test_loader = torch.utils.data.DataLoader(test_dataset, batch_size=BATCH_SIZE, shuffle=False, collate_fn=lambda b: collate_fn(b, PAD_IDX))

    for epoch in range(N_EPOCHS):
        model.train()
        epoch_loss = 0
        total_batches = len(train_loader)
        for batch_idx, batch in enumerate(train_loader):
            ko_padded, ko_lens, zh_padded, zh_lens = batch
            src = ko_padded.to(device).transpose(0, 1).contiguous()
            trg = zh_padded.to(device).transpose(0, 1).contiguous()
            optimizer.zero_grad()
            output = model(src, ko_lens, trg)
            output_dim = output.shape[-1]
            output = output[1:].view(-1, output_dim)
            trg = trg[1:].view(-1)
            loss = criterion(output, trg)
            loss.backward()
            torch.nn.utils.clip_grad_norm_(model.parameters(), CLIP)
            optimizer.step()
            epoch_loss += loss.item()
            progress = (batch_idx + 1) / total_batches * 100
            print(f'Epoch: {epoch+1:02}, Batch: {batch_idx+1}/{total_batches}, Progress: {progress:.2f}%, Train Loss: {loss.item():.3f}', end='\r')
        train_loss = epoch_loss / total_batches
        train_losses.append(train_loss)
        print(f'\nEpoch: {epoch+1:02}, Train Loss: {train_loss:.3f}')

        model.eval()
        test_loss = 0
        test_batches = 0
        with torch.no_grad():
            for batch in test_loader:
                ko_padded, ko_lens, zh_padded, zh_lens = batch
                src = ko_padded.to(device).transpose(0, 1).contiguous()
                trg = zh_padded.to(device).transpose(0, 1).contiguous()
                output = model(src, ko_lens, trg)
                output_dim = output.shape[-1]
                output = output[1:].view(-1, output_dim)
                trg = trg[1:].view(-1)
                loss = criterion(output, trg)
                test_loss += loss.item()
                test_batches += 1
        test_loss = test_loss / test_batches
        test_losses.append(test_loss)
        print(f'Epoch: {epoch+1:02}, Test Loss: {test_loss:.3f}')

        scheduler.step(test_loss)

        if test_loss < best_test_loss:
            best_test_loss = test_loss
            patience_counter = 0
            best_model_state = {k: v.cpu().clone() for k, v in model.state_dict().items()}
            print(f'  -> 测试损失下降，保存最佳模型状态！')
            
            # 实时保存到磁盘，方便训练中途测试
            model_folder = 'Translate Model'
            if not os.path.exists(model_folder):
                os.makedirs(model_folder)
            
            # 使用固定名称保存“最新最佳”模型，方便测试脚本读取
            torch.save(best_model_state, os.path.join(model_folder, 'best_model_latest.pth'))
            with open(os.path.join(model_folder, 'korean_vocab.pkl'), 'wb') as f:
                pickle.dump(korean_vocab, f)
            with open(os.path.join(model_folder, 'chinese_vocab.pkl'), 'wb') as f:
                pickle.dump(chinese_vocab, f)
            print(f'  -> 已同步保存最新最佳模型和词汇表到 {model_folder} 文件夹')
        else:
            patience_counter += 1
            print(f'  -> 测试损失未下降 ({patience_counter}/{EARLY_STOPPING_PATIENCE})')
            if patience_counter >= EARLY_STOPPING_PATIENCE:
                print(f'\n早停触发！连续{EARLY_STOPPING_PATIENCE}个epoch测试损失未下降，停止训练。')
                break

    if best_model_state is not None:
        model.load_state_dict(best_model_state)
        print(f'已恢复最佳模型，测试损失: {best_test_loss:.3f}')

    model.eval()
    with torch.no_grad():
        for ko_t, zh_t in zip(ko_test_tensor, zh_test_tensor):
            src = ko_t.unsqueeze(1).to(device)
            src_lens = [src.shape[0]]
            trg = zh_t.unsqueeze(1).to(device)
            output = model(src, src_lens, trg, teacher_forcing_ratio=0)
            output = output.argmax(2)
            pass

    model_folder = 'Translate Model'
    if not os.path.exists(model_folder):
        os.makedirs(model_folder)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    model_path = os.path.join(model_folder, f'model_{timestamp}.pth')
    torch.save(model.state_dict(), model_path)
    print(f'Model saved to {model_path}')

    vocab_path = os.path.join(model_folder, f'korean_vocab_{timestamp}.pkl')
    with open(vocab_path, 'wb') as f:
        pickle.dump(korean_vocab, f)
    vocab_path = os.path.join(model_folder, f'chinese_vocab_{timestamp}.pkl')
    with open(vocab_path, 'wb') as f:
        pickle.dump(chinese_vocab, f)

    print('Train Losses:', train_losses)
    print('Test Losses:', test_losses)
    print(f'Best Test Loss: {best_test_loss:.3f}')

# 运行交互界面
create_gui()